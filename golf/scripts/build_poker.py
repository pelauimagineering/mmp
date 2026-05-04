#!/usr/bin/env python3
"""
build_poker.py — Convert the MMP Poker xlsx into data/poker.json.

Drop the latest xlsx into the project root (or pass it as an argument)
and run:  python3 scripts/build_poker.py

The script:
  - reads every Poker_YYYY_YY sheet
  - parses each "Event: <Month Year> (<Host>)" block
  - normalizes tie/runner-up cells (handles stray whitespace)
  - flags placeholder events as played: false
  - derives finalStandings from the last played event of each season
"""
from __future__ import annotations
import json, re, sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUTS = sorted(ROOT.glob("MMP Poker*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
OUT = ROOT / "data" / "poker.json"

EVENT_RE = re.compile(r"^Event:\s*(?P<month>\w+)\s+(?P<year>\d{4})\s*(?:\((?:Host:\s*)?(?P<host>[^)]*)\))?\s*$", re.I)
MONTH_ORDER = ["Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct"]


def clean_num(v):
    if v is None: return None
    if isinstance(v, str):
        s = v.strip()
        if not s: return None
        try: return float(s)
        except ValueError: return None
    return float(v)


def parse_event_header(text: str):
    m = EVENT_RE.match(text.strip())
    if not m: return None
    month_raw = m.group("month")[:3].title()  # "March" -> "Mar"
    year = int(m.group("year"))
    host = (m.group("host") or "").strip()
    if host.lower() in ("no games", ""): host_clean = None
    else: host_clean = host
    skipped = (m.group("host") or "").strip().lower() == "no games"
    return {"month": month_raw, "year": year, "host": host_clean, "skipped": skipped}


def parse_sheet(ws) -> list[dict]:
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    events = []
    i = 0
    while i < len(rows):
        cell = rows[i][1] if len(rows[i]) > 1 else None
        if isinstance(cell, str) and cell.lstrip().startswith("Event:"):
            hdr = parse_event_header(cell)
            if not hdr:
                i += 1; continue
            event = {**hdr, "results": [], "totals": None}
            if hdr["skipped"]:
                events.append(event); i += 1; continue
            # next row is column header; skip it. Player rows follow until "Total".
            j = i + 1
            if j < len(rows) and (rows[j][1] or "") == "Name":
                j += 1
            while j < len(rows):
                r = rows[j]
                name = r[1]
                if not isinstance(name, str): break
                if name == "Total":
                    event["totals"] = {
                        "appearance": clean_num(r[2]), "win": clean_num(r[3]),
                        "tie": clean_num(r[4]), "runnerUp": clean_num(r[5]),
                        "eventTotal": clean_num(r[6]), "ytdTotal": clean_num(r[7]),
                    }
                    j += 1; break
                if name.lstrip().startswith("Event:"): break
                event["results"].append({
                    "name": name,
                    "appearance": clean_num(r[2]),
                    "win": clean_num(r[3]),
                    "tie": clean_num(r[4]),
                    "runnerUp": clean_num(r[5]),
                    "eventTotal": clean_num(r[6]) or 0.0,
                    "ytdTotal": clean_num(r[7]) or 0.0,
                    "rank": clean_num(r[8]),
                })
                j += 1
            # Determine if event was actually played
            played = any((p.get("appearance") or 0) > 0 or (p.get("win") or 0) > 0 or
                         (p.get("tie") or 0) > 0 or (p.get("runnerUp") or 0) > 0
                         for p in event["results"])
            event["played"] = played
            event["id_suffix"] = f"{hdr['month'].lower()}-{hdr['year']}"
            events.append(event)
            i = j
        else:
            i += 1
    return events


def build_season(sheet_name: str, events: list[dict]) -> dict:
    # season id: Poker_2025_26 -> "2025-26"
    m = re.match(r"Poker_(\d{4})_(\d{2})", sheet_name)
    sid = f"{m.group(1)}-{m.group(2)}" if m else sheet_name
    label = sid.replace("-", "–")  # en-dash for display
    out_events = []
    skipped_months = []
    for e in events:
        if e.get("skipped"):
            skipped_months.append(f"{e['month']} {e['year']}")
            continue
        out_events.append({
            "id": f"{sid}-{e['month'].lower()}",
            "month": e["month"], "year": e["year"], "host": e["host"],
            "played": e["played"],
            "results": e["results"] if e["played"] else [],
            "totals": e["totals"],
        })

    # Final standings = ranking by ytd from last played event
    last_played = next((e for e in reversed(out_events) if e["played"]), None)
    standings = []
    if last_played:
        sorted_results = sorted(last_played["results"], key=lambda p: -p["ytdTotal"])
        rank = 0; prev = None
        for idx, p in enumerate(sorted_results, start=1):
            if p["ytdTotal"] != prev:
                rank = idx; prev = p["ytdTotal"]
            standings.append({"rank": rank, "name": p["name"], "total": p["ytdTotal"]})

    return {
        "id": sid, "label": label,
        "events": out_events,
        "skippedMonths": skipped_months,
        "finalStandings": standings,
    }


def main(argv):
    src = Path(argv[1]) if len(argv) > 1 else (DEFAULT_INPUTS[0] if DEFAULT_INPUTS else None)
    if not src or not src.exists():
        print("Usage: build_poker.py <path-to-xlsx>", file=sys.stderr); return 1
    print(f"Reading {src.name}")
    wb = openpyxl.load_workbook(src, data_only=True)
    sheets = [n for n in wb.sheetnames if n.startswith("Poker_")]
    sheets.sort(reverse=True)  # newest season first
    seasons = [build_season(n, parse_sheet(wb[n])) for n in sheets]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({"sport": "poker", "source": src.name, "seasons": seasons}, indent=2))
    # Summary
    for s in seasons:
        played = sum(1 for e in s["events"] if e["played"])
        winner = s["finalStandings"][0]["name"] if s["finalStandings"] else "—"
        winner_pts = s["finalStandings"][0]["total"] if s["finalStandings"] else 0
        print(f"  {s['label']}: {played} events played · winner {winner} ({winner_pts})")
    print(f"Wrote {OUT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
